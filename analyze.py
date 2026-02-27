import requests
import json
import ssl
from datetime import datetime, timedelta
import time
import traceback
import pandas as pd
import concurrent.futures

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
    'Connection': 'keep-alive',
    'Referer': 'https://www.tpex.org.tw/'
}

def validate_trading_day(date_str):
    """
    使用體積極小的 '市場成交概況' API 來快速預檢當天是否為有效交易日。
    這比直接抓整份法人買賣超 (T86) 輕量得多，適合用來做前置測試。
    """
    # MI_INDEX type=MS 是市場成交概況，回傳資料極少
    url = f"https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date={date_str}&type=MS"
    data = get_json(url)
    # 如果 data['stat'] 為 'OK'，代表當天有交易紀錄
    return data and data.get('stat') == 'OK'

def fetch_twse(date="20260223"):
    t86_url = f"https://www.twse.com.tw/fund/T86?response=json&date={date}&selectType=ALL"
    mi_url = f"https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date={date}&type=ALLBUT0999"
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f_t86 = executor.submit(get_json, t86_url)
        f_mi = executor.submit(get_json, mi_url)
        t86_data = f_t86.result()
        mi_data = f_mi.result()
    
    if not t86_data or 'data' not in t86_data:
        print("Failed to get TWSE T86")
        return []
        
    prices = {}
    if mi_data and 'tables' in mi_data:
        # MI_INDEX tables structure, usually the 9th table is closing prices
        target_table = None
        for table in mi_data['tables']:
            title = table.get('title', '')
            if '每日收盤行情' in title:
                target_table = table
                break
        
        if target_table and 'fields' in target_table:
            mi_fields = target_table['fields']
            try:
                idx_mi_code = mi_fields.index('證券代號')
                idx_mi_close = mi_fields.index('收盤價')
                idx_mi_vol = mi_fields.index('成交股數')
                idx_mi_val = mi_fields.index('成交金額')
                
                for row in target_table['data']:
                    code = row[idx_mi_code].strip()
                    price_str = row[idx_mi_close].replace(',', '')
                    vol_str = row[idx_mi_vol].replace(',', '')
                    val_str = row[idx_mi_val].replace(',', '')
                    
                    try:
                        close_p = float(price_str)
                    except ValueError:
                        close_p = 0.0
                        
                    vwap = close_p
                    if vol_str.isdigit() and val_str.isdigit():
                        vol = int(vol_str)
                        val = int(val_str)
                        if vol > 0:
                            vwap = val / vol
                    
                    prices[code] = {'close': close_p, 'vwap': vwap}
            except Exception as e:
                print("Error parsing TWSE MI_INDEX fields:", e)
    
    results = []
    t86_fields = t86_data['fields']
    idx_code = t86_fields.index('證券代號')
    idx_name = t86_fields.index('證券名稱')
    
    try:
        idx_foreign = next(i for i, f in enumerate(t86_fields) if '外陸資買賣超股數(不含外資自營商)' in f)
    except StopIteration:
        idx_foreign = next(i for i, f in enumerate(t86_fields) if '外資' in f and '買賣超' in f)
        
    try:
        idx_it = next(i for i, f in enumerate(t86_fields) if f == '投信買賣超股數')
    except StopIteration:
        idx_it = next(i for i, f in enumerate(t86_fields) if '投信' in f and '買賣超' in f)

    for row in t86_data['data']:
        code = row[idx_code].strip()
        name = row[idx_name].strip()
        
        # If no price or price is 0, we can calculate value
        if code not in prices or prices[code]['close'] == 0:
            continue
            
        # 排除 ETF 與非普通股 (代號長度非 4 或 0 開頭)
        if len(code) != 4 or code.startswith('0'):
            continue
            
        try:
            foreign_shares = int(row[idx_foreign].replace(',', ''))
            it_shares = int(row[idx_it].replace(',', ''))
        except ValueError:
            continue
            
        foreign_value = foreign_shares * prices[code]['vwap']
        it_value = it_shares * prices[code]['vwap']
        
        results.append({
            'market': 'TWSE',
            'code': code,
            'name': name,
            'price': prices[code]['close'],
            'vwap': prices[code]['vwap'],
            'foreign_val': foreign_value,
            'it_val': it_value,
            'foreign_shares': foreign_shares,
            'it_shares': it_shares
        })
        
    return results

def fetch_tpex(date_roc="115/02/23"):
    # tpex T86 equivalent
    t86_url = f"https://www.tpex.org.tw/web/stock/3insti/daily_trade/3itrade_hedge_result.php?l=zh-tw&se=EW&t=D&d={date_roc}"
    # tpex MI_INDEX equivalent
    mi_url = f"https://www.tpex.org.tw/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php?l=zh-tw&d={date_roc}"
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f_t86 = executor.submit(get_json, t86_url)
        f_mi = executor.submit(get_json, mi_url)
        t86_data = f_t86.result()
        mi_data = f_mi.result()
    
    results = []
    if not t86_data or 'tables' not in t86_data or not t86_data['tables']:
        print("Failed to get TPEX T86")
        return results
        
    if not mi_data or 'tables' not in mi_data or not mi_data['tables']:
        print("Failed to get TPEX closing prices")
        return results
        
    prices = {}
    mi_table = mi_data['tables'][0]
    for row in mi_table.get('data', []):
        code = row[0].strip()
        price_str = row[2].replace(',', '') # idx 2 is '收盤'
        try:
            close_p = float(price_str)
        except ValueError:
            close_p = 0.0
            
        vwap = close_p
        try:
            if len(row) > 9:
                vol_str = str(row[8]).replace(',', '')
                val_str = str(row[9]).replace(',', '')
                if vol_str.isdigit() and val_str.isdigit() and int(vol_str) > 0:
                    vwap = int(val_str) / int(vol_str)
                elif str(row[7]).replace('.', '').replace(',', '').isdigit():
                    vwap = float(str(row[7]).replace(',', ''))
        except Exception:
            pass
            
        prices[code] = {'close': close_p, 'vwap': vwap}
            
    t86_table = t86_data['tables'][0]
    for row in t86_table.get('data', []):
        # TPEX format: 0=代號, 1=名稱
        # 4=外資買賣超, 7=外資自營買賣超, 10=外資合計, 13=投信買賣超
        code = str(row[0]).strip()
        name = str(row[1]).strip()
        if code not in prices or prices[code]['close'] == 0:
            continue
            
        # 排除 ETF 與非普通股 (代號長度非 4 或 0 開頭)
        if len(code) != 4 or code.startswith('0'):
            continue
            
        try:
            foreign_shares = int(str(row[4]).replace(',', ''))
            it_shares = int(str(row[13]).replace(',', ''))
        except (ValueError, IndexError):
            continue
            
        foreign_value = foreign_shares * prices[code]['vwap']
        it_value = it_shares * prices[code]['vwap']
        
        results.append({
            'market': 'TPEX',
            'code': code,
            'name': name,
            'price': prices[code]['close'],
            'vwap': prices[code]['vwap'],
            'foreign_val': foreign_value,
            'it_val': it_value,
            'foreign_shares': foreign_shares,
            'it_shares': it_shares
        })
        
    return results

def format_val(val):
    if val >= 0:
        return f"+{val/100000000:.2f}億元"
    else:
        return f"{val/100000000:.2f}億元"

def analyze(target_date_str=None):
    if not target_date_str:
        target_date_str = datetime.now().strftime('%Y%m%d')
        
    year = int(target_date_str[:4])
    month = target_date_str[4:6]
    day = target_date_str[6:8]
    roc_year = year - 1911
    
    twse_date = target_date_str
    tpex_date = f"{roc_year:03d}/{month}/{day}"
    
    print(f"Fetching data from TWSE ({twse_date}) and TPEx ({tpex_date}) in parallel...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f_twse = executor.submit(fetch_twse, twse_date)
        f_tpex = executor.submit(fetch_tpex, tpex_date)
        twse_data = f_twse.result()
        tpex_data = f_tpex.result()
    
    all_data = twse_data + tpex_data
    if not all_data:
        print(f"No data for {target_date_str}. The market might be closed.")
        return False # Return False instead of raising, to let the loop handle it
    
    # ... rest of analysis logic ...
    # (Note: Need to make sure all_data logic can continue or return status)
    # Let's keep it simple: if all_data exists, it returns True at the end of function
    # Wait, I see analyze function doesn't return anything. I'll modify it to return success status.

        
    print(f"Successfully processed {len(all_data)} stocks.")
    print("="*60)
    
    # Sort by foreign value
    foreign_buy = sorted([d for d in all_data if d['foreign_val'] > 0], key=lambda x: x['foreign_val'], reverse=True)
    foreign_sell = sorted([d for d in all_data if d['foreign_val'] < 0], key=lambda x: x['foreign_val'])
    
    # Sort by IT value
    it_buy = sorted([d for d in all_data if d['it_val'] > 0], key=lambda x: x['it_val'], reverse=True)
    it_sell = sorted([d for d in all_data if d['it_val'] < 0], key=lambda x: x['it_val'])
    
    print("\n### 外資買超排名 (依成交值)")
    for i, d in enumerate(foreign_buy[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['foreign_val'])}")
        
    print("\n### 外資賣超排名 (依成交值)")
    for i, d in enumerate(foreign_sell[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['foreign_val'])}")
        
    print("\n### 投信買超排名 (依成交值)")
    for i, d in enumerate(it_buy[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['it_val'])}")
        
    print("\n### 投信賣超排名 (依成交值)")
    for i, d in enumerate(it_sell[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['it_val'])}")

    print("\n" + "="*60)
    
    # 同向與反向分析
    # 同向買超: 外資買超 > 0 且 投信買超 > 0, 依加總值排序
    same_buy = [d for d in all_data if d['foreign_val'] > 0 and d['it_val'] > 0]
    same_buy.sort(key=lambda x: x['foreign_val'] + x['it_val'], reverse=True)
    
    print("\n### 土洋同買超 (外資與投信皆買超，依總買超金額排序)")
    for i, d in enumerate(same_buy[:10], 1):
        total = d['foreign_val'] + d['it_val']
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 總計 {format_val(total)} (外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])})")
        
    # 同向賣超: 外資賣超 < 0 且 投信賣超 < 0
    same_sell = [d for d in all_data if d['foreign_val'] < 0 and d['it_val'] < 0]
    same_sell.sort(key=lambda x: x['foreign_val'] + x['it_val'])
    
    print("\n### 土洋同賣超 (外資與投信皆賣超，依總賣超金額排序)")
    for i, d in enumerate(same_sell[:10], 1):
        total = d['foreign_val'] + d['it_val']
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 總計 {format_val(total)} (外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])})")
        
    print("\n" + "="*60)
    
    # 土洋對作: 外資與投信方向相反
    # 分為: 外資買/投信賣, 外資賣/投信買 (依兩者絕對值加總排序表示激烈程度)
    opp_fb_is = [d for d in all_data if d['foreign_val'] > 0 > d['it_val']]
    opp_fb_is.sort(key=lambda x: abs(x['foreign_val']) + abs(x['it_val']), reverse=True)
    
    print("\n### 土洋對作: 外資買超、投信賣超 (依對作規模排序)")
    for i, d in enumerate(opp_fb_is[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])}")
        
    opp_fs_ib = [d for d in all_data if d['foreign_val'] < 0 < d['it_val']]
    opp_fs_ib.sort(key=lambda x: abs(x['foreign_val']) + abs(x['it_val']), reverse=True)
    
    print("\n### 土洋對作: 外資賣超、投信買超 (依對作規模排序)")
    for i, d in enumerate(opp_fs_ib[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])}")
        
    print("\n" + "="*60 + "\n完成！")

    # 輸出成四欄位、分上市櫃的 Excel 報表
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        import subprocess
        import sys
        print("首次執行，正在安裝 openpyxl 套件...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # 移除預設工作表
        
        # 建立格式與字體
        report_date = f"{year}/{month}/{day}"
        
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        dark_red_fill = PatternFill(start_color="FF8080", end_color="FF8080", fill_type="solid")
        light_green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        dark_green_fill = PatternFill(start_color="80FF80", end_color="80FF80", fill_type="solid")
        
        date_font = Font(name='微軟正黑體', size=12, bold=True, color="000000")
        header_font = Font(name='微軟正黑體', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        sub_header_font = Font(name='微軟正黑體', size=11, bold=True)
        base_font = Font(name='微軟正黑體', size=11)
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        for market_key, sheet_name in [('TWSE', '上市'), ('TPEX', '上櫃')]:
            ws = wb.create_sheet(title=sheet_name)
            
            # 第一列: 日期
            ws.append([f"{report_date}"])
            ws.cell(row=1, column=1).font = date_font
            ws.cell(row=1, column=1).alignment = left_align
            
            # 篩選出該市場的資料
            market_data = [d for d in all_data if d['market'] == market_key]
            
            # 依買賣超金額排序 (由大到小 / 由深到淺即負數由小到大)
            fb = sorted([d for d in market_data if d['foreign_shares'] > 0], key=lambda x: x['foreign_val'], reverse=True)
            fs = sorted([d for d in market_data if d['foreign_shares'] < 0], key=lambda x: x['foreign_val'])
            ib = sorted([d for d in market_data if d['it_shares'] > 0], key=lambda x: x['it_val'], reverse=True)
            isell = sorted([d for d in market_data if d['it_shares'] < 0], key=lambda x: x['it_val'])
            
            max_rows = max(len(fb), len(fs), len(ib), len(isell))
            
            # 第二列: 大標題
            row2 = [
                "外資買超", "", "", "", "", "", "",
                "外資賣超", "", "", "", "", "", "",
                "投信買超", "", "", "", "", "", "",
                "投信賣超", "", "", "", "", ""
            ]
            ws.append(row2)
            
            # 第三列: 子標題
            sub_headers = [
                "證券代號", "證券名稱", "收盤價", "均價", "股數", "估價(百萬)", "",
                "證券代號", "證券名稱", "收盤價", "均價", "股數", "估價(百萬)", "",
                "證券代號", "證券名稱", "收盤價", "均價", "股數", "估價(百萬)", "",
                "證券代號", "證券名稱", "收盤價", "均價", "股數", "估價(百萬)"
            ]
            ws.append(sub_headers)
            
            # 合併第二列儲存格
            ws.merge_cells("A2:F2")
            ws.merge_cells("H2:M2")
            ws.merge_cells("O2:T2")
            ws.merge_cells("V2:AA2")
            
            # 設定前三列樣式
            for cell in ws[2]:
                cell.font = header_font
                cell.alignment = center_align
                if cell.value:
                    cell.fill = header_fill
            for cell in ws[3]:
                cell.font = sub_header_font
                cell.alignment = center_align
                if cell.value:
                    cell.fill = sub_header_fill
            
            # 計算每檔股票的狀態 (同向或反向)
            stock_state = {}
            for d in market_data:
                f_val, i_val = d['foreign_shares'], d['it_shares']
                if (f_val > 0 and i_val > 0) or (f_val < 0 and i_val < 0):
                    if abs(f_val) > abs(i_val):
                        stock_state[d['code']] = (dark_red_fill, light_red_fill)
                    else:
                        stock_state[d['code']] = (light_red_fill, dark_red_fill)
                elif (f_val > 0 and i_val < 0) or (f_val < 0 and i_val > 0):
                    if abs(f_val) > abs(i_val):
                        stock_state[d['code']] = (dark_green_fill, light_green_fill)
                    else:
                        stock_state[d['code']] = (light_green_fill, dark_green_fill)
                else:
                    stock_state[d['code']] = (None, None)

            # 輔助函式：取得股票資料與對應的顏色
            def get_stock_data(lst, idx, val_key, shares_key, is_foreign):
                if idx < len(lst):
                    st = lst[idx]
                    fills = stock_state.get(st['code'], (None, None))
                    fill = fills[0] if is_foreign else fills[1]
                    code_val = int(st['code']) if st['code'].isdigit() else st['code']
                    return [code_val, st['name'], st['price'], st['vwap'], st[shares_key], st[val_key] / 1000000], fill
                return ["", "", "", "", "", ""], None

            # 寫入各分類的排名資料
            for row_i in range(max_rows):
                fb_data, fb_fill = get_stock_data(fb, row_i, 'foreign_val', 'foreign_shares', True)
                fs_data, fs_fill = get_stock_data(fs, row_i, 'foreign_val', 'foreign_shares', True)
                ib_data, ib_fill = get_stock_data(ib, row_i, 'it_val', 'it_shares', False)
                is_data, is_fill = get_stock_data(isell, row_i, 'it_val', 'it_shares', False)
                
                row_idx = row_i + 4 # 標題佔 3 列
                
                col_settings = [
                    (1, fb_data[0], None, None), (2, fb_data[1], fb_fill, None), (3, fb_data[2], None, '#,##0.00'), (4, fb_data[3], None, '#,##0.00'), (5, fb_data[4], None, '#,##0'), (6, fb_data[5], None, '#,##0.00'),
                    (8, fs_data[0], None, None), (9, fs_data[1], fs_fill, None), (10, fs_data[2], None, '#,##0.00'), (11, fs_data[3], None, '#,##0.00'), (12, fs_data[4], None, '#,##0'), (13, fs_data[5], None, '#,##0.00'),
                    (15, ib_data[0], None, None), (16, ib_data[1], ib_fill, None), (17, ib_data[2], None, '#,##0.00'), (18, ib_data[3], None, '#,##0.00'), (19, ib_data[4], None, '#,##0'), (20, ib_data[5], None, '#,##0.00'),
                    (22, is_data[0], None, None), (23, is_data[1], is_fill, None), (24, is_data[2], None, '#,##0.00'), (25, is_data[3], None, '#,##0.00'), (26, is_data[4], None, '#,##0'), (27, is_data[5], None, '#,##0.00')
                ]
                
                for col, val, fill, num_fmt in col_settings:
                    if val != "":
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.font = base_font
                        cell.alignment = right_align if isinstance(val, (int, float)) else center_align
                        if fill and col in [2, 9, 16, 23]: # Only highlight the Name column
                            cell.fill = fill
                        if num_fmt and isinstance(val, (int, float)):
                            cell.number_format = num_fmt

            # 調整欄位寬度
            for c in ['A', 'H', 'O', 'V']:
                ws.column_dimensions[c].width = 10
            for c in ['B', 'I', 'P', 'W']:
                ws.column_dimensions[c].width = 12
            for c in ['C', 'J', 'Q', 'X']:
                ws.column_dimensions[c].width = 10
            for c in ['D', 'K', 'R', 'Y']:
                ws.column_dimensions[c].width = 11
            for c in ['E', 'L', 'S', 'Z']:
                ws.column_dimensions[c].width = 16
            for c in ['F', 'M', 'T', 'AA']:
                ws.column_dimensions[c].width = 13
            for c in ['G', 'N', 'U']:
                ws.column_dimensions[c].width = 3

        filename = f"market_analysis_{target_date_str}.xlsx"
        wb.save(filename)
        print(f"\n已成功輸出多欄位變色 Excel 報表: {filename}")
    except Exception as e:
        print(f"\n輸出報表時發生錯誤: {e}")

if __name__ == '__main__':
    import sys
    input_date = sys.argv[1] if len(sys.argv) > 1 else None
    
    # 無論是有輸入日期還是自動觸發，如果發現當天沒開盤，都應該回溯尋找
    success = False
    
    # 決定起始日期
    if input_date:
        # 使用者指定的日期 (格式 YYYYMMDD)
        start_date_obj = datetime.strptime(input_date, '%Y%m%d')
        print(f"User requested analysis starting from: {input_date}")
    else:
        # 自動模式，從今天開始找
        start_date_obj = datetime.now()
        print(f"Automatic daily trigger starting from today...")

    # 智慧回溯循環 (最多往回找 10 天交易日)
    for i in range(10):
        current_date_str = (start_date_obj - timedelta(days=i)).strftime('%Y%m%d')
        print(f"--- [快速預檢] 測試日期: {current_date_str} (Day {i+1}) ---")
        
        if validate_trading_day(current_date_str):
            print(f"✅ 成功命中有效交易日: {current_date_str}！ 準備開始執行重型分析任務...")
            try:
                analyze(current_date_str)
                success = True
                break
            except Exception as e:
                print(f"❌ 執行分析時發生非預期錯誤: {e}")
                traceback.print_exc()
                # 即使預檢成功，分析失敗也應該結束，避免無限回溯
                break
        else:
            print(f"⚠️ 日期 {current_date_str} 休市中，自動跳過...")
            continue
    
    if not success:
        print("🚨 任務失敗：在最近的 10 天內找不到任何開盤紀錄，請檢查證交所連線或網站狀態。")
        sys.exit(1)
